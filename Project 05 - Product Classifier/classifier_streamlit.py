import streamlit as st
import pandas as pd
import pickle
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Load models
with open('beauty_category.pk1', 'rb') as f:
    loaded_model = pickle.load(f)
with open('beauty_sub_category.pk1', 'rb') as f:
    loaded_model_2 = pickle.load(f)
with open('manufacturer.pk1', 'rb') as f:
    loaded_model_3 = pickle.load(f)
with open('manufacturer_brand.pk1', 'rb') as f:
    loaded_model_4 = pickle.load(f)
with open('manufacturer_sub-brand.pk1', 'rb') as f:
    loaded_model_5 = pickle.load(f)
with open('manufacturer_franchise.pk1', 'rb') as f:
    loaded_model_6 = pickle.load(f)

# Streamlit app
st.title('Product Classifier')

# Instructions
st.markdown("""
<div style='text-align: center; font-size: 22px;'>
    Dear user please make sure your columns are named <span style="color: green;">ProductNew</span> and <span style="color: green;">EAN</span>.
</div>
""", unsafe_allow_html=True)

# Upload file
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file is not None:
    # Read the Excel file
    new_products = pd.read_excel(uploaded_file)

    # Ensure the necessary columns are present
    if 'ProductNew' in new_products.columns and 'EAN' in new_products.columns:
        X_new = new_products['ProductNew']

        # Classify categories and manufacturer hierarchy
        new_products['Category'] = loaded_model.predict(X_new)
        new_products['SubCategory'] = loaded_model_2.predict(X_new)
        new_products['Manufacturer'] = loaded_model_3.predict(X_new)
        new_products['ManufacturerBrand'] = loaded_model_4.predict(X_new)
        new_products['ManufacturerSub-Brand'] = loaded_model_5.predict(X_new)
        new_products['ManufacturerFranchise'] = loaded_model_6.predict(X_new)

        # Replace 'Not applicable' with blanks in ManufacturerFranchise where necessary
        new_products['ManufacturerFranchise'].replace('Not applicable', '', inplace=True)

        # Convert EAN to string and remove commas
        new_products['EAN'] = new_products['EAN'].apply(lambda x: str(x).replace(',', ''))

        # Display the updated data with classification
        st.write('### Classified Products')
        st.dataframe(new_products)

        # Convert the DataFrame to an Excel file in memory and adjust column widths
        def to_excel(df):
            output = BytesIO()
            writer = pd.ExcelWriter(output, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            writer.close()

            # Adjust column widths
            workbook = load_workbook(output)
            worksheet = workbook.active

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            adjusted_output = BytesIO()
            workbook.save(adjusted_output)
            return adjusted_output.getvalue()

        excel_data = to_excel(new_products)
        st.download_button(label="Download Classified Data",
                           data=excel_data,
                           file_name='New_Products_Classified.xlsx')

    else:
        st.error("The uploaded file does not contain the required 'ProductNew' and 'EAN' columns.")
